import pandas as pd
import openpyxl

# Read all sheets from the Excel file
def read_excel_sheets(file_path):
    # Use sheet_name=None to load all sheets
    xls = pd.read_excel(file_path, sheet_name=None)
    
    # Create an empty list to hold dataframes
    dataframes = []
    
    # Iterate over each sheet in the Excel file 
    for sheet_name, df in xls.items():
        # Add a new column for the sheet name
        df['SheetName'] = sheet_name
        # Append the dataframe to the list
        dataframes.append(df)
    
    # Concatenate all dataframes into one
    combined_df = pd.concat(dataframes, ignore_index=True)
    
    return combined_df

# Usage example


def read_excel_openpxyl(file_paht):
    colors = {}
    num=0
    wb = openpyxl.load_workbook(file_paht,data_only=True)
    for sheet in wb:
        fs = sheet
        # print(sheet.title)
        fs_count_row = fs.max_row 
        fs_count_col = fs.max_column 
        for row in range(1,fs_count_row+1):
            for column in range(1,2):
            # for column in range(1,fs_count_col+1):
                cell_color = fs.cell(column=column, row=row)
                cell_value = fs.cell(column=column, row=row).value
                bgColor = cell_color.fill.bgColor.index
                fgColor = cell_color.fill.fgColor.index
                # if (bgColor=='00000000') or (fgColor=='00000000'):
                #     continue
                # else:
                    # print("Background color index of cell (",cell_value, ") is", bgColor)
                colors[cell_value] = {"sheet": str(fs), "bg": bgColor, "fg" :fgColor}
                    # print("Background color index of cell (",row,column, ") is", bgColor)
                    # print("Foreground color index of cell (",row,column, ") is", fgColor)

    print(len(colors.keys()))
    return colors

def map_color(color_name):
    return color_name


color_dict = {
    'FFD9EAD3' : 'green', 
    'FFFFF2CC' : 'yellow', 
    'FFF4CCCC' : 'red', 
    'FFB4A7D6' : 'purple', 
    'FFFCE5CD' : 'orange',
    '00000000' : 'blank', 
    'FFCFE2F3'  : 'blue', 
    'FFD9D2E9' : 'pink'
}

if __name__ == "__main__":
    # file_path = 'picso_orig.xlsx'
    file_path = 'nov5.xlsx'
    combined_df = read_excel_sheets(file_path)
    colors_pd = read_excel_openpxyl(file_path)



    combined_df["bgColor"] = combined_df["message_id"].map(lambda x: map_color(colors_pd[x]["bg"]) if x in colors_pd else "00000000")

    # print(combined_df.loc[combined_df["SheetName"] == "9-15-24"])
    combined_df.to_csv('./nov5_combined_with_colors.csv', index=False, encoding='utf-8-sig')